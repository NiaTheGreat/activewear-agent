"""Main agent orchestrator coordinating all workflow phases."""

from pathlib import Path
from typing import Dict, List, Optional, Set, Tuple

from openpyxl import load_workbook
from rich.console import Console
from rich.panel import Panel

from agent.state import AgentState
from config import OUTPUT_DIR, settings
from models.criteria import SearchCriteria
from models.manufacturer import ContactInfo, Manufacturer
from tools.criteria_collector import CriteriaCollector
from tools.data_extractor import DataExtractor
from tools.evaluator import Evaluator
from tools.excel_generator import ExcelGenerator
from tools.query_generator import QueryGenerator
from tools.web_scraper import WebScraper
from tools.web_searcher import WebSearcher
from utils.llm import get_client

console = Console()


class ManufacturerResearchAgent:
    """Main agent that orchestrates the manufacturer research workflow."""

    def __init__(self):
        """Initialize the agent and all tools."""
        self.state = AgentState.INIT

        # Initialize all tools
        self.criteria_collector = CriteriaCollector()
        self.query_generator = QueryGenerator()
        self.web_searcher = WebSearcher()
        self.web_scraper = WebScraper()
        self.data_extractor = DataExtractor()
        self.evaluator = Evaluator()
        self.excel_generator = ExcelGenerator()

        # Data storage
        self.criteria: Optional[SearchCriteria] = None
        self.queries: List[str] = []
        self.urls: List[str] = []
        self.scraped_data: dict = {}
        self.manufacturers: List[Manufacturer] = []
        self.output_path: Optional[Path] = None
        self.max_manufacturers: int = settings.MAX_MANUFACTURERS  # Can be overridden by user input

    def run(self) -> Path:
        """
        Run the complete manufacturer research workflow.

        Returns:
            Path to generated Excel report

        Raises:
            Exception: If any critical error occurs during execution
        """
        try:
            # Phase 1: Collect Criteria
            self.state = AgentState.COLLECTING_CRITERIA
            self.criteria = self._collect_criteria()

            # Ask how many manufacturers to research
            self._ask_manufacturer_count()

            # Ask if user wants to skip web search
            skip_search = console.input(
                "\n[bold]Skip web search and enter URLs manually? (y/N):[/bold] "
            ).lower()

            if skip_search == "y":
                # Skip to manual URL input
                console.print("\n[cyan]Skipping web search...[/cyan]\n")
                self.queries = []
                manual_urls = self.web_searcher.manual_input(skip_prompt=True, max_count=self.max_manufacturers)
                # Filter manual URLs against existing database
                self.urls = self._filter_new_urls(manual_urls) if manual_urls else []
            else:
                # Phase 2: Generate Search Queries
                self.state = AgentState.GENERATING_QUERIES
                self.queries = self._generate_queries()

                # Phase 3: Search Web
                self.state = AgentState.SEARCHING
                self.urls = self._search_web()

            # Filter URLs to avoid re-scraping existing ones
            if self.urls:
                self.urls = self._filter_new_urls(self.urls)

            # Phase 4: Scrape Websites
            self.state = AgentState.SCRAPING
            self.scraped_data = self._scrape_websites()

            # Phase 5: Extract Data
            # (Note: No separate state, part of scraping phase)
            self.manufacturers = self._extract_data()

            # Phase 6: Evaluate Manufacturers
            self.state = AgentState.EVALUATING
            self.manufacturers = self._evaluate_manufacturers()

            # Phase 7: Generate Output
            self.state = AgentState.OUTPUTTING
            self.output_path = self._generate_output()

            # Complete
            self.state = AgentState.COMPLETE
            self._display_summary()

            return self.output_path

        except KeyboardInterrupt:
            console.print("\n\n[yellow]Agent interrupted by user.[/yellow]\n")
            self.state = AgentState.ERROR
            raise

        except Exception as e:
            console.print(f"\n[red]Error during execution:[/red] {e}\n")
            self.state = AgentState.ERROR
            raise

    def _collect_criteria(self) -> SearchCriteria:
        """
        Phase 1: Collect search criteria from user.

        Returns:
            SearchCriteria object
        """
        return self.criteria_collector.collect()

    def _ask_manufacturer_count(self) -> None:
        """Ask user how many manufacturers to research."""
        console.print(
            f"\n[dim]Default: {settings.MAX_MANUFACTURERS} manufacturers[/dim]"
        )

        while True:
            count_input = console.input(
                f"[bold]How many manufacturers to research? (1-50, Enter for {settings.MAX_MANUFACTURERS}):[/bold] "
            ).strip()

            # Use default if empty
            if not count_input:
                self.max_manufacturers = settings.MAX_MANUFACTURERS
                console.print(
                    f"[green]✓ Researching {self.max_manufacturers} manufacturers[/green]\n"
                )
                break

            # Validate input
            try:
                count = int(count_input)
                if 1 <= count <= 50:
                    self.max_manufacturers = count
                    console.print(
                        f"[green]✓ Researching {self.max_manufacturers} manufacturers[/green]\n"
                    )
                    break
                else:
                    console.print(
                        "[yellow]Please enter a number between 1 and 50.[/yellow]"
                    )
            except ValueError:
                console.print("[yellow]Please enter a valid number.[/yellow]")

    def _generate_queries(self) -> List[str]:
        """
        Phase 2: Generate search queries from criteria.

        Returns:
            List of search query strings
        """
        console.print(
            f"\n[bold cyan]Step 2: Generating Search Queries[/bold cyan]\n"
        )

        queries = self.query_generator.generate(self.criteria)

        console.print(f"[green]✓ Generated {len(queries)} search queries:[/green]")
        for i, query in enumerate(queries, 1):
            console.print(f"  {i}. [dim]{query}[/dim]")

        console.print()
        return queries

    def _search_web(self) -> List[str]:
        """
        Phase 3: Search web for manufacturer URLs.

        Returns:
            List of manufacturer website URLs
        """
        urls = self.web_searcher.search(self.queries, max_urls=self.max_manufacturers)

        # If search failed, offer manual input
        if not urls:
            urls = self.web_searcher.manual_input(max_count=self.max_manufacturers)

        if not urls:
            console.print(
                "[yellow]Warning: No URLs provided. Cannot continue.[/yellow]\n"
            )

        return urls

    def _filter_new_urls(self, urls: List[str]) -> List[str]:
        """
        Filter URLs to only include those not already in manufacturers_scores.xlsx.
        This prevents re-scraping URLs that have already been processed.

        Args:
            urls: List of URLs from search results

        Returns:
            Filtered list of URLs not in the cumulative file
        """
        cumulative_filename = "manufacturers_scores.xlsx"
        cumulative_path = OUTPUT_DIR / cumulative_filename

        # If cumulative file doesn't exist, all URLs are new
        if not cumulative_path.exists():
            console.print("[dim]No existing manufacturers file found. All URLs are new.[/dim]\n")
            return urls

        try:
            # Read existing URLs from cumulative file
            console.print(f"\n[cyan]Checking for existing URLs in {cumulative_filename}...[/cyan]")

            wb_existing = load_workbook(cumulative_path, read_only=True)
            ws_existing = wb_existing.active

            existing_urls: Set[str] = set()

            # Extract existing URLs from column O (Source URL = column 15)
            for row in range(2, ws_existing.max_row + 1):
                url_cell = ws_existing.cell(row=row, column=15)
                if url_cell.value:
                    existing_urls.add(str(url_cell.value).strip())

            wb_existing.close()

            console.print(f"[dim]Found {len(existing_urls)} existing URLs in database[/dim]")

            # Filter to only new URLs
            new_urls = [url for url in urls if url not in existing_urls]

            filtered_count = len(urls) - len(new_urls)

            if filtered_count > 0:
                console.print(
                    f"[yellow]✓ Filtered out {filtered_count} already-scraped URLs[/yellow]"
                )

            if new_urls:
                console.print(
                    f"[green]✓ {len(new_urls)} new URLs to scrape[/green]\n"
                )
            else:
                console.print(
                    "[yellow]⚠️  All URLs already exist in database. No new URLs to scrape.[/yellow]\n"
                )

            return new_urls

        except Exception as e:
            console.print(
                f"[yellow]Warning: Could not read cumulative file: {e}[/yellow]"
            )
            console.print("[dim]Proceeding with all URLs...[/dim]\n")
            return urls

    def _scrape_websites(self) -> dict:
        """
        Phase 4: Scrape manufacturer websites.

        Returns:
            Dictionary mapping URLs to scraped content
        """
        if not self.urls:
            console.print("[yellow]No URLs to scrape. Skipping...[/yellow]\n")
            return {}

        # Limit to user-specified number of URLs
        urls_to_scrape = self.urls[: self.max_manufacturers]

        return self.web_scraper.scrape_urls(urls_to_scrape)

    def _extract_data(self) -> List[Manufacturer]:
        """
        Phase 5: Extract structured data from scraped content.

        Returns:
            List of Manufacturer objects
        """
        if not self.scraped_data:
            console.print(
                "[yellow]No scraped data available. Cannot extract manufacturers.[/yellow]\n"
            )
            return []

        return self.data_extractor.extract(self.scraped_data)

    def _evaluate_manufacturers(self) -> List[Manufacturer]:
        """
        Phase 6: Evaluate and score manufacturers.

        Returns:
            List of evaluated Manufacturer objects (sorted by score)
        """
        if not self.manufacturers:
            console.print(
                "[yellow]No manufacturers to evaluate.[/yellow]\n"
            )
            return []

        return self.evaluator.evaluate(self.manufacturers, self.criteria)

    def _generate_output(self) -> Path:
        """
        Phase 7: Generate Excel output.

        Returns:
            Path to Excel file
        """
        if not self.manufacturers:
            console.print(
                "[yellow]No manufacturers to output. Creating empty report.[/yellow]\n"
            )

        excel_path = self.excel_generator.generate(self.manufacturers)

        # Generate failures report if there are any failures
        scrape_failures = self.web_scraper.failed_urls
        extraction_failures = self.data_extractor.failed_extractions

        if scrape_failures or extraction_failures:
            failures_path = self.excel_generator.generate_failures_report(
                scrape_failures, extraction_failures
            )

        return excel_path

    def rescore(self) -> Path:
        """
        Re-score existing manufacturers from the cumulative Excel file.

        Reads manufacturer data from manufacturers_scores.xlsx, lets user
        pick criteria (preset), re-evaluates with the current scoring algorithm,
        and rewrites the Excel with updated scores.

        Returns:
            Path to the updated Excel file
        """
        try:
            self.state = AgentState.RESCORING

            console.print(
                "\n[bold cyan]Rescore Mode[/bold cyan] - Re-evaluate existing manufacturers\n"
            )

            # Step 1: Collect or load criteria
            console.print("[bold]First, select the criteria to score against:[/bold]\n")
            self.criteria = self._collect_criteria()

            # Step 2: Read manufacturers from Excel
            manufacturers, date_added_map = self._read_manufacturers_from_excel()

            if not manufacturers:
                console.print("[yellow]No manufacturers found in Excel. Nothing to rescore.[/yellow]\n")
                self.state = AgentState.COMPLETE
                return OUTPUT_DIR / "manufacturers_scores.xlsx"

            console.print(
                f"[green]Loaded {len(manufacturers)} manufacturers from Excel[/green]\n"
            )

            # Step 3: Re-evaluate
            self.state = AgentState.EVALUATING
            self.manufacturers = self.evaluator.evaluate(manufacturers, self.criteria)

            # Step 4: Rewrite Excel
            self.state = AgentState.OUTPUTTING
            self.output_path = self.excel_generator.rewrite_scores(
                self.manufacturers, date_added_map
            )

            # Complete
            self.state = AgentState.COMPLETE
            self._display_rescore_summary()

            return self.output_path

        except KeyboardInterrupt:
            console.print("\n\n[yellow]Rescore interrupted by user.[/yellow]\n")
            self.state = AgentState.ERROR
            raise

        except Exception as e:
            console.print(f"\n[red]Error during rescore:[/red] {e}\n")
            self.state = AgentState.ERROR
            raise

    def _read_manufacturers_from_excel(self) -> Tuple[List[Manufacturer], Dict[str, str]]:
        """
        Read manufacturer data from the cumulative Excel file and reconstruct
        Manufacturer objects.

        Returns:
            Tuple of (list of Manufacturer objects, dict mapping source_url to date_added)
        """
        cumulative_path = OUTPUT_DIR / "manufacturers_scores.xlsx"

        if not cumulative_path.exists():
            console.print("[yellow]No manufacturers_scores.xlsx found.[/yellow]\n")
            return [], {}

        console.print(f"[cyan]Reading manufacturers from {cumulative_path}...[/cyan]")

        wb = load_workbook(cumulative_path, read_only=True)
        ws = wb.active

        manufacturers = []
        date_added_map: Dict[str, str] = {}

        for row in range(2, ws.max_row + 1):
            # Read all columns
            name = ws.cell(row=row, column=2).value
            if not name:
                continue  # Skip empty rows

            location = ws.cell(row=row, column=3).value
            website = ws.cell(row=row, column=4).value or ""
            moq_raw = ws.cell(row=row, column=5).value
            materials_raw = ws.cell(row=row, column=8).value
            certs_raw = ws.cell(row=row, column=9).value
            methods_raw = ws.cell(row=row, column=10).value
            email = ws.cell(row=row, column=11).value
            phone = ws.cell(row=row, column=12).value
            address = ws.cell(row=row, column=13).value
            source_url = ws.cell(row=row, column=15).value or website
            date_added = ws.cell(row=row, column=16).value

            # Parse MOQ
            moq = None
            if moq_raw is not None and moq_raw != "Unknown":
                try:
                    moq = int(moq_raw)
                except (ValueError, TypeError):
                    pass

            # Parse comma-separated lists
            def parse_list(raw) -> List[str]:
                if not raw or raw in ("Unknown", "None listed"):
                    return []
                return [item.strip() for item in str(raw).split(",") if item.strip()]

            materials = parse_list(materials_raw)
            certifications = parse_list(certs_raw)
            production_methods = parse_list(methods_raw)

            # Clean contact fields
            clean_email = email if email and email != "Unknown" else None
            clean_phone = phone if phone and phone != "Unknown" else None
            clean_address = address if address and address != "Unknown" else None
            clean_location = location if location and location != "Unknown" else None

            try:
                manufacturer = Manufacturer(
                    name=str(name),
                    website=str(website) if website else str(source_url),
                    location=clean_location,
                    contact=ContactInfo(
                        email=clean_email,
                        phone=clean_phone,
                        address=clean_address,
                    ),
                    materials=materials,
                    production_methods=production_methods,
                    moq=moq,
                    certifications=certifications,
                    source_url=str(source_url),
                )
                manufacturers.append(manufacturer)

                # Preserve date added
                if source_url and date_added:
                    date_added_map[str(source_url)] = str(date_added)

            except Exception as e:
                console.print(
                    f"  [yellow]Skipping row {row} ({name}): {e}[/yellow]"
                )

        wb.close()

        console.print(
            f"[green]Read {len(manufacturers)} manufacturers from Excel[/green]\n"
        )

        return manufacturers, date_added_map

    def _display_rescore_summary(self) -> None:
        """Display summary after rescoring."""
        summary_lines = [
            f"[bold]Manufacturers Rescored:[/bold] {len(self.manufacturers)}",
        ]

        if self.manufacturers:
            summary_lines.append(
                f"[bold]Top Match:[/bold] {self.manufacturers[0].name} "
                f"([green]{self.manufacturers[0].match_score}[/green])"
            )

            # Score distribution
            high = sum(1 for m in self.manufacturers if m.match_score >= 70)
            mid = sum(1 for m in self.manufacturers if 50 <= m.match_score < 70)
            low = sum(1 for m in self.manufacturers if m.match_score < 50)

            summary_lines.append("")
            summary_lines.append("[bold cyan]--- Score Distribution ---[/bold cyan]")
            summary_lines.append(f"  [green]70+:[/green] {high} manufacturers")
            summary_lines.append(f"  [yellow]50-69:[/yellow] {mid} manufacturers")
            summary_lines.append(f"  [red]<50:[/red] {low} manufacturers")

        summary_lines.append(f"\n[bold]Report Location:[/bold] {self.output_path}")
        summary_lines.append("[dim]No API calls were made (local rescore only)[/dim]")

        console.print(
            Panel(
                "\n".join(summary_lines),
                title="[bold green]Rescore Complete[/bold green]",
                border_style="green",
                padding=(1, 2),
            )
        )
        console.print()

    def _display_summary(self) -> None:
        """Display final summary of results."""
        summary_lines = [
            f"[bold]Manufacturers Found:[/bold] {len(self.manufacturers)}",
            f"[bold]Top Match:[/bold] {self.manufacturers[0].name if self.manufacturers else 'None'} "
            f"([green]{self.manufacturers[0].match_score}[/green])"
            if self.manufacturers
            else "[bold]Top Match:[/bold] None",
            f"[bold]Report Location:[/bold] {self.output_path}",
        ]

        # Add failure info if any
        scrape_failures_count = len(self.web_scraper.failed_urls)
        extraction_failures_count = len(self.data_extractor.failed_extractions)
        total_failures = scrape_failures_count + extraction_failures_count

        if total_failures > 0:
            failure_details = []
            if scrape_failures_count > 0:
                failure_details.append(f"{scrape_failures_count} scraping")
            if extraction_failures_count > 0:
                failure_details.append(f"{extraction_failures_count} extraction")

            summary_lines.append(
                f"[yellow]Failures:[/yellow] {total_failures} total ({', '.join(failure_details)}) - see failures Excel"
            )

        # Add cost information
        claude_client = get_client()
        usage_stats = claude_client.get_usage_stats()

        summary_lines.append("")  # Blank line separator
        summary_lines.append("[bold cyan]--- API Usage & Costs ---[/bold cyan]")

        # Claude API costs
        summary_lines.append(
            f"[bold]Claude API:[/bold] {usage_stats['total_tokens']:,} tokens "
            f"({usage_stats['input_tokens']:,} in / {usage_stats['output_tokens']:,} out)"
        )
        summary_lines.append(
            f"[bold]Claude Cost:[/bold] [green]${usage_stats['total_cost']:.4f}[/green]"
        )

        # Brave Search API costs
        search_queries = len(self.queries) if self.queries else 0
        summary_lines.append(f"[bold]Brave Searches:[/bold] {search_queries} queries (free tier: 2,000/month)")

        # Total estimated cost
        total_cost = usage_stats['total_cost']
        summary_lines.append(f"[bold]Total Cost:[/bold] [green]${total_cost:.4f}[/green]")

        summary_text = "\n".join(summary_lines)

        console.print(
            Panel(
                summary_text,
                title="[bold green]✓ Research Complete[/bold green]",
                border_style="green",
                padding=(1, 2),
            )
        )
        console.print()
