"""Generate formatted Excel reports from manufacturer data."""

from datetime import datetime
from pathlib import Path
from typing import List, Optional
from zoneinfo import ZoneInfo

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from rich.console import Console

from config import OUTPUT_DIR, settings
from models.manufacturer import Manufacturer

console = Console()


class ExcelGenerator:
    """Generates formatted Excel reports of manufacturer evaluations."""

    def __init__(self):
        """Initialize the Excel generator."""
        self.problematic_urls = []  # Track URLs that can't be added to Excel

    @staticmethod
    def _sanitize_for_excel(value) -> str:
        """
        Sanitize a value for safe Excel export by removing problematic characters.
        Used only for failure reports where we want to preserve data even if imperfect.

        Args:
            value: Value to sanitize

        Returns:
            Sanitized string safe for Excel
        """
        if value is None:
            return ""

        text = str(value)

        # Remove control characters that Excel can't handle
        import re
        # Remove control characters (0x00-0x1F except tab, newline, carriage return)
        sanitized = re.sub(r'[\x00-\x08\x0B-\x0C\x0E-\x1F\x7F-\x9F]', '', text)

        return sanitized

    @staticmethod
    def _has_problematic_characters(value) -> bool:
        """
        Check if a value contains characters that can't be used in Excel.

        Args:
            value: Value to check

        Returns:
            True if value has problematic characters, False otherwise
        """
        if value is None or value == "":
            return False

        text = str(value)

        # Check for control characters that Excel can't handle
        import re
        # Control characters (0x00-0x1F except tab, newline, carriage return, and 0x7F-0x9F)
        problematic = re.search(r'[\x00-\x08\x0B-\x0C\x0E-\x1F\x7F-\x9F]', text)

        return problematic is not None

    def _save_data_quality_urls(self, timestamp: str) -> Path:
        """
        Save URLs with data quality issues for manual research.

        Args:
            timestamp: Timestamp string for filename

        Returns:
            Path to the saved file
        """
        filename = f"data_quality_issues_{timestamp}.txt"
        filepath = OUTPUT_DIR / filename

        with open(filepath, "w", encoding="utf-8") as f:
            f.write("DATA QUALITY ISSUES - Manual Research Required\n")
            f.write("=" * 70 + "\n")
            f.write(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
            f.write(f"Total URLs: {len(self.problematic_urls)}\n")
            f.write("=" * 70 + "\n\n")
            f.write("These manufacturers had data with problematic characters.\n")
            f.write("They were excluded from the main report to maintain data quality.\n")
            f.write("Research these URLs manually to extract information.\n\n")
            f.write("=" * 70 + "\n\n")

            for url in self.problematic_urls:
                f.write(f"{url}\n")

        console.print(
            f"[yellow]⚠️  Data quality issues saved to:[/yellow] {filepath}\n"
        )

        return filepath

    def _update_cumulative_file(self, clean_manufacturers: List[Manufacturer], timestamp: str) -> Optional[Path]:
        """
        Update or create the cumulative manufacturers_scores.xlsx file.
        Only adds manufacturers with new/unique URLs.

        Args:
            clean_manufacturers: List of clean Manufacturer objects from this run
            timestamp: Timestamp string for logging

        Returns:
            Path to cumulative file, or None if no new manufacturers added
        """
        cumulative_filename = "manufacturers_scores.xlsx"
        cumulative_path = OUTPUT_DIR / cumulative_filename

        # Read existing URLs if file exists
        existing_urls = set()
        next_row = 2  # Start after header

        if cumulative_path.exists():
            try:
                # Load existing workbook
                wb_existing = load_workbook(cumulative_path)
                ws_existing = wb_existing.active

                # Extract existing URLs from column O (Source URL)
                for row in range(2, ws_existing.max_row + 1):
                    url_cell = ws_existing.cell(row=row, column=15)  # Column O = 15
                    if url_cell.value:
                        existing_urls.add(url_cell.value)

                next_row = ws_existing.max_row + 1
                wb_existing.close()

                console.print(f"[dim]Found {len(existing_urls)} existing manufacturers in cumulative file[/dim]")

            except Exception as e:
                console.print(f"[yellow]Warning: Could not read existing cumulative file: {e}[/yellow]")
                # Continue with empty set, will overwrite file

        # Filter to only new manufacturers (by URL)
        new_manufacturers = []
        for manufacturer in clean_manufacturers:
            if manufacturer.source_url not in existing_urls:
                new_manufacturers.append(manufacturer)

        if not new_manufacturers and cumulative_path.exists():
            console.print(f"[dim]No new manufacturers to add to cumulative file (all URLs already exist)[/dim]\n")
            return cumulative_path

        # Create or append to cumulative file
        if cumulative_path.exists() and existing_urls:
            # Append to existing file
            wb = load_workbook(cumulative_path)
            ws = wb.active

            # Check if "Date Added" column exists, add it if missing
            header_row_value = ws.cell(row=1, column=16).value
            if header_row_value != "Date Added":
                # Add "Date Added" header
                ws.cell(row=1, column=16, value="Date Added")
                ws.cell(row=1, column=16).font = Font(bold=True, color="FFFFFF")
                ws.cell(row=1, column=16).fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                ws.cell(row=1, column=16).alignment = Alignment(horizontal="center", vertical="center")
                ws.column_dimensions["P"].width = 20

                # Fill empty "Date Added" for existing rows
                for row in range(2, next_row):
                    ws.cell(row=row, column=16, value="[Before tracking]")

            console.print(f"[cyan]Adding {len(new_manufacturers)} new manufacturers to cumulative file...[/cyan]")

            # Get current timestamp for new manufacturers (Central Time)
            central_tz = ZoneInfo("America/Chicago")
            date_added = datetime.now(central_tz).strftime("%Y-%m-%d %H:%M %Z")

            # Append new manufacturers
            for manufacturer in new_manufacturers:
                row_data = manufacturer.to_excel_row()

                # Calculate new rank based on current row
                rank = next_row - 1

                # Write each cell
                ws.cell(row=next_row, column=1, value=rank)
                ws.cell(row=next_row, column=2, value=row_data["Name"])
                ws.cell(row=next_row, column=3, value=row_data["Location"])
                ws.cell(row=next_row, column=4, value=row_data["Website"])
                ws.cell(row=next_row, column=5, value=row_data["MOQ"])
                ws.cell(row=next_row, column=6, value=row_data["Match Score"])
                ws.cell(row=next_row, column=7, value=row_data["Confidence"])
                ws.cell(row=next_row, column=8, value=row_data["Materials"])
                ws.cell(row=next_row, column=9, value=row_data["Certifications"])
                ws.cell(row=next_row, column=10, value=row_data["Production Methods"])
                ws.cell(row=next_row, column=11, value=row_data["Email"])
                ws.cell(row=next_row, column=12, value=row_data["Phone"])
                ws.cell(row=next_row, column=13, value=row_data["Address"])
                ws.cell(row=next_row, column=14, value=row_data["Notes"])
                ws.cell(row=next_row, column=15, value=row_data["Source URL"])
                ws.cell(row=next_row, column=16, value=date_added)

                # Color code match scores
                score_cell = ws.cell(row=next_row, column=6)
                score = row_data["Match Score"]

                if score >= 70:
                    score_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                elif score >= 50:
                    score_cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                else:
                    score_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

                # Wrap text for notes
                ws.cell(row=next_row, column=14).alignment = Alignment(wrap_text=True)

                next_row += 1

        else:
            # Create new cumulative file (same structure as timestamped file)
            console.print(f"[cyan]Creating new cumulative file with {len(new_manufacturers)} manufacturers...[/cyan]")

            # Get current timestamp for new manufacturers (Central Time)
            central_tz = ZoneInfo("America/Chicago")
            date_added = datetime.now(central_tz).strftime("%Y-%m-%d %H:%M %Z")

            wb = Workbook()
            ws = wb.active
            ws.title = "Manufacturers"

            # Define column headers
            headers = [
                "Rank",
                "Name",
                "Location",
                "Website",
                "MOQ",
                "Match Score",
                "Confidence",
                "Materials",
                "Certifications",
                "Production Methods",
                "Email",
                "Phone",
                "Address",
                "Notes",
                "Source URL",
                "Date Added",
            ]

            # Write headers with formatting
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num)
                cell.value = header
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")

            # Write data rows
            for idx, manufacturer in enumerate(new_manufacturers, 1):
                row_data = manufacturer.to_excel_row()
                row_data["Rank"] = idx

                row_num = idx + 1

                # Write each cell
                ws.cell(row=row_num, column=1, value=row_data["Rank"])
                ws.cell(row=row_num, column=2, value=row_data["Name"])
                ws.cell(row=row_num, column=3, value=row_data["Location"])
                ws.cell(row=row_num, column=4, value=row_data["Website"])
                ws.cell(row=row_num, column=5, value=row_data["MOQ"])
                ws.cell(row=row_num, column=6, value=row_data["Match Score"])
                ws.cell(row=row_num, column=7, value=row_data["Confidence"])
                ws.cell(row=row_num, column=8, value=row_data["Materials"])
                ws.cell(row=row_num, column=9, value=row_data["Certifications"])
                ws.cell(row=row_num, column=10, value=row_data["Production Methods"])
                ws.cell(row=row_num, column=11, value=row_data["Email"])
                ws.cell(row=row_num, column=12, value=row_data["Phone"])
                ws.cell(row=row_num, column=13, value=row_data["Address"])
                ws.cell(row=row_num, column=14, value=row_data["Notes"])
                ws.cell(row=row_num, column=15, value=row_data["Source URL"])
                ws.cell(row=row_num, column=16, value=date_added)

                # Color code match scores
                score_cell = ws.cell(row=row_num, column=6)
                score = row_data["Match Score"]

                if score >= 70:
                    score_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                elif score >= 50:
                    score_cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                else:
                    score_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

                # Wrap text for notes
                ws.cell(row=row_num, column=14).alignment = Alignment(wrap_text=True)

            # Adjust column widths
            column_widths = {
                "A": 6,   # Rank
                "B": 25,  # Name
                "C": 20,  # Location
                "D": 35,  # Website
                "E": 10,  # MOQ
                "F": 12,  # Match Score
                "G": 12,  # Confidence
                "H": 30,  # Materials
                "I": 30,  # Certifications
                "J": 30,  # Production Methods
                "K": 25,  # Email
                "L": 15,  # Phone
                "M": 30,  # Address
                "N": 50,  # Notes
                "O": 40,  # Source URL
                "P": 20,  # Date Added (with timezone)
            }

            for col, width in column_widths.items():
                ws.column_dimensions[col].width = width

            # Freeze header row
            ws.freeze_panes = "A2"

        # Save cumulative workbook
        wb.save(cumulative_path)

        total_manufacturers = len(existing_urls) + len(new_manufacturers)
        console.print(
            f"[green]✓ Cumulative file updated:[/green] {cumulative_path} "
            f"[dim]({total_manufacturers} total manufacturers, {len(new_manufacturers)} new)[/dim]\n"
        )

        return cumulative_path

    def _sync_to_notion(self, clean_manufacturers: List[Manufacturer]) -> None:
        """
        Sync manufacturers to Notion database (optional).

        Args:
            clean_manufacturers: List of clean Manufacturer objects from this run
        """
        if not settings.NOTION_ENABLED:
            return

        try:
            from tools.notion_uploader import NotionUploader

            uploader = NotionUploader()
            if uploader.is_enabled():
                uploader.sync_manufacturers(clean_manufacturers)
        except ImportError:
            console.print(
                "[yellow]⚠️  Notion integration requires: pip install notion-client[/yellow]\n"
            )
        except Exception as e:
            console.print(f"[yellow]⚠️  Notion sync error: {e}[/yellow]\n")

    def generate(self, manufacturers: List[Manufacturer]) -> Path:
        """
        Generate Excel file with manufacturer data.
        Only creates/updates the cumulative manufacturers_scores.xlsx file.

        Args:
            manufacturers: List of evaluated Manufacturer objects (should be sorted)

        Returns:
            Path to cumulative Excel file
        """
        # Reset problematic URLs for this generation
        self.problematic_urls = []

        console.print(
            f"\n[bold cyan]Step 7: Processing Manufacturers[/bold cyan] ({len(manufacturers)} manufacturers)\n"
        )

        # Filter out manufacturers with data quality issues
        clean_manufacturers = []
        for manufacturer in manufacturers:
            row_data = manufacturer.to_excel_row()

            # Check if any field has problematic characters
            has_problems = False
            for key, value in row_data.items():
                if key in ["Rank", "Match Score"]:  # Skip numeric fields
                    continue
                if self._has_problematic_characters(value):
                    has_problems = True
                    break

            if has_problems:
                # Track this URL for manual research
                source_url = row_data.get("Source URL", manufacturer.source_url)
                if source_url:
                    self.problematic_urls.append(source_url)
                console.print(
                    f"  [dim]Skipping manufacturer with data quality issues: {row_data.get('Name', 'Unknown')}[/dim]"
                )
            else:
                clean_manufacturers.append(manufacturer)

        # Generate timestamp for data quality issues file
        timestamp = datetime.now().strftime(settings.TIMESTAMP_FORMAT)

        # Save data quality issues to separate file if any
        if self.problematic_urls:
            self._save_data_quality_urls(timestamp)

        # Update cumulative manufacturers_scores.xlsx file
        cumulative_path = self._update_cumulative_file(clean_manufacturers, timestamp)

        # Sync to Notion (optional)
        self._sync_to_notion(clean_manufacturers)

        return cumulative_path

    def generate_failures_report(
        self, scrape_failures: list, extraction_failures: list
    ) -> Optional[Path]:
        """
        Generate Excel file with all failures (scraping + extraction).

        Args:
            scrape_failures: List of (url, reason) tuples from scraping
            extraction_failures: List of (url, reason) tuples from extraction

        Returns:
            Path to generated Excel file, or None if no failures
        """
        # Combine all failures
        all_failures = []

        for url, reason in scrape_failures:
            all_failures.append(
                {
                    "URL": url,
                    "Failure Type": "Scraping Failed",
                    "Error Reason": reason,
                    "Status": "Not Scraped",
                }
            )

        for url, reason in extraction_failures:
            all_failures.append(
                {
                    "URL": url,
                    "Failure Type": "Extraction Failed",
                    "Error Reason": reason,
                    "Status": "Scraped but not extracted",
                }
            )

        if not all_failures:
            return None

        console.print(
            f"\n[yellow]Generating failures report ({len(all_failures)} failures)...[/yellow]\n"
        )

        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Failed URLs"

        # Define column headers
        headers = ["#", "URL", "Failure Type", "Status", "Error Reason"]

        # Write headers with formatting
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(
                start_color="C00000", end_color="C00000", fill_type="solid"
            )
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Write data rows with sanitization
        for idx, failure in enumerate(all_failures, 1):
            ws.cell(row=idx + 1, column=1, value=idx)
            ws.cell(row=idx + 1, column=2, value=self._sanitize_for_excel(failure["URL"]))
            ws.cell(row=idx + 1, column=3, value=self._sanitize_for_excel(failure["Failure Type"]))
            ws.cell(row=idx + 1, column=4, value=self._sanitize_for_excel(failure["Status"]))
            ws.cell(row=idx + 1, column=5, value=self._sanitize_for_excel(failure["Error Reason"]))

            # Color code by failure type
            type_cell = ws.cell(row=idx + 1, column=3)
            if failure["Failure Type"] == "Scraping Failed":
                type_cell.fill = PatternFill(
                    start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"
                )
            else:  # Extraction Failed
                type_cell.fill = PatternFill(
                    start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"
                )

            # Wrap text for error reason
            ws.cell(row=idx + 1, column=5).alignment = Alignment(wrap_text=True)

        # Adjust column widths
        ws.column_dimensions["A"].width = 5  # #
        ws.column_dimensions["B"].width = 45  # URL
        ws.column_dimensions["C"].width = 18  # Failure Type
        ws.column_dimensions["D"].width = 25  # Status
        ws.column_dimensions["E"].width = 60  # Error Reason

        # Freeze header row
        ws.freeze_panes = "A2"

        # Generate filename with timestamp
        timestamp = datetime.now().strftime(settings.TIMESTAMP_FORMAT)
        filename = f"failures_{timestamp}.xlsx"
        output_path = OUTPUT_DIR / filename

        # Save workbook
        wb.save(output_path)

        console.print(f"[yellow]✓ Failures report saved to:[/yellow] {output_path}\n")

        return output_path

    def rewrite_scores(
        self,
        manufacturers: List[Manufacturer],
        date_added_map: Optional[dict] = None,
    ) -> Path:
        """
        Rewrite the cumulative manufacturers_scores.xlsx with re-scored data.

        Overwrites the file with updated match_score, confidence, and notes
        while preserving the original Date Added values.

        Args:
            manufacturers: List of re-evaluated Manufacturer objects (sorted by score)
            date_added_map: Dict mapping source_url -> original date_added string

        Returns:
            Path to the rewritten Excel file
        """
        cumulative_path = OUTPUT_DIR / "manufacturers_scores.xlsx"
        date_added_map = date_added_map or {}

        console.print(
            f"\n[bold cyan]Rewriting Scores[/bold cyan] ({len(manufacturers)} manufacturers)\n"
        )

        # Create fresh workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Manufacturers"

        # Column headers (same layout as original)
        headers = [
            "Rank", "Name", "Location", "Website", "MOQ",
            "Match Score", "Confidence", "Materials", "Certifications",
            "Production Methods", "Email", "Phone", "Address",
            "Notes", "Source URL", "Date Added",
        ]

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(
                start_color="366092", end_color="366092", fill_type="solid"
            )
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Write data rows
        central_tz = ZoneInfo("America/Chicago")
        rescore_timestamp = datetime.now(central_tz).strftime("%Y-%m-%d %H:%M %Z")

        for idx, manufacturer in enumerate(manufacturers, 1):
            row_data = manufacturer.to_excel_row()
            row_num = idx + 1

            # Preserve original date_added or mark as rescored
            original_date = date_added_map.get(manufacturer.source_url, "")
            if original_date:
                date_value = f"{original_date} (rescored {rescore_timestamp})"
            else:
                date_value = f"Rescored {rescore_timestamp}"

            ws.cell(row=row_num, column=1, value=idx)
            ws.cell(row=row_num, column=2, value=row_data["Name"])
            ws.cell(row=row_num, column=3, value=row_data["Location"])
            ws.cell(row=row_num, column=4, value=row_data["Website"])
            ws.cell(row=row_num, column=5, value=row_data["MOQ"])
            ws.cell(row=row_num, column=6, value=row_data["Match Score"])
            ws.cell(row=row_num, column=7, value=row_data["Confidence"])
            ws.cell(row=row_num, column=8, value=row_data["Materials"])
            ws.cell(row=row_num, column=9, value=row_data["Certifications"])
            ws.cell(row=row_num, column=10, value=row_data["Production Methods"])
            ws.cell(row=row_num, column=11, value=row_data["Email"])
            ws.cell(row=row_num, column=12, value=row_data["Phone"])
            ws.cell(row=row_num, column=13, value=row_data["Address"])
            ws.cell(row=row_num, column=14, value=row_data["Notes"])
            ws.cell(row=row_num, column=15, value=row_data["Source URL"])
            ws.cell(row=row_num, column=16, value=date_value)

            # Color code match scores
            score_cell = ws.cell(row=row_num, column=6)
            score = row_data["Match Score"]

            if score >= 70:
                score_cell.fill = PatternFill(
                    start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"
                )
            elif score >= 50:
                score_cell.fill = PatternFill(
                    start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"
                )
            else:
                score_cell.fill = PatternFill(
                    start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"
                )

            # Wrap text for notes
            ws.cell(row=row_num, column=14).alignment = Alignment(wrap_text=True)

        # Adjust column widths
        column_widths = {
            "A": 6, "B": 25, "C": 20, "D": 35, "E": 10,
            "F": 12, "G": 12, "H": 30, "I": 30, "J": 30,
            "K": 25, "L": 15, "M": 30, "N": 50, "O": 40, "P": 25,
        }
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width

        # Freeze header row
        ws.freeze_panes = "A2"

        # Save
        wb.save(cumulative_path)

        console.print(
            f"[green]Rescored {len(manufacturers)} manufacturers -> {cumulative_path}[/green]\n"
        )

        return cumulative_path
